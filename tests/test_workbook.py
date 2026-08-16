"""Tests for workbook writing.

The organisers' validator fails silently from our side: a workbook that looks
correct in Excel can still be rejected at upload because a cell holds a string
instead of a number. These tests encode the validator's rules so a regression is
caught here rather than at 17:55 on the day.
"""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from src.config import PATHS, find_company, load_companies
from src.workbook import (
    WorkbookError,
    find_header_row,
    verify_workbook,
    write_company_workbook,
)


class WriteWorkbookTests(unittest.TestCase):
    def setUp(self):
        self._temp = tempfile.TemporaryDirectory()
        self.out = Path(self._temp.name)
        self.company = find_company("HD")
        self.values = {
            "Net sales": 45210.0,
            "Adjusted diluted EPS": 4.72,
            "Comparable sales, total company": 1.4,
        }

    def tearDown(self):
        self._temp.cleanup()

    def _write(self, values=None):
        return write_company_workbook(self.company, values or self.values, output_dir=self.out)

    def test_writes_the_expected_filename(self):
        path, _ = self._write()
        self.assertEqual(path.name, "HD-FY2026Q2.xlsx")

    def test_writes_into_the_period_column_below_the_header(self):
        path, written = self._write()
        self.assertEqual([cell.cell for cell in written], ["C7", "C8", "C9"])

    def test_values_are_stored_as_numbers_not_strings(self):
        path, _ = self._write()
        sheet = load_workbook(path)["Summary"]
        for row in (7, 8, 9):
            value = sheet.cell(row=row, column=3).value
            self.assertIsInstance(value, (int, float), msg=f"row {row} is {type(value).__name__}")

    def test_template_structure_is_preserved(self):
        path, _ = self._write()
        workbook = load_workbook(path)
        self.assertIn("Summary", workbook.sheetnames)
        self.assertIn("Instructions", workbook.sheetnames)
        sheet = workbook["Summary"]
        self.assertEqual(sheet.cell(row=6, column=1).value, "Metric")
        self.assertEqual(sheet.cell(row=6, column=3).value, "FY2026Q2")

    def test_metric_labels_and_units_are_untouched(self):
        path, _ = self._write()
        sheet = load_workbook(path)["Summary"]
        for offset, metric in enumerate(self.company.metrics, start=1):
            self.assertEqual(sheet.cell(row=6 + offset, column=1).value, metric.label)
            self.assertEqual(sheet.cell(row=6 + offset, column=2).value, metric.units)

    def test_string_forecast_is_coerced_to_a_number(self):
        path, written = self._write({**self.values, "Net sales": "45210"})
        self.assertEqual(written[0].value, 45210.0)
        self.assertIsInstance(load_workbook(path)["Summary"]["C7"].value, (int, float))

    def test_missing_metric_is_rejected(self):
        incomplete = dict(self.values)
        del incomplete["Adjusted diluted EPS"]
        with self.assertRaises(WorkbookError) as caught:
            self._write(incomplete)
        self.assertIn("Adjusted diluted EPS", str(caught.exception))

    def test_non_numeric_forecast_is_rejected(self):
        with self.assertRaises(WorkbookError):
            self._write({**self.values, "Net sales": "not a number"})

    def test_infinite_forecast_is_rejected(self):
        with self.assertRaises(WorkbookError):
            self._write({**self.values, "Net sales": float("inf")})

    def test_boolean_forecast_is_rejected(self):
        # bool is an int subclass; the validator would accept it as a number.
        with self.assertRaises(WorkbookError):
            self._write({**self.values, "Net sales": True})

    def test_negative_percentages_are_allowed(self):
        _, written = self._write({**self.values, "Comparable sales, total company": -2.3})
        self.assertEqual(written[2].value, -2.3)


class VerifyWorkbookTests(unittest.TestCase):
    def setUp(self):
        self._temp = tempfile.TemporaryDirectory()
        self.out = Path(self._temp.name)

    def tearDown(self):
        self._temp.cleanup()

    def test_every_company_round_trips(self):
        for company in load_companies():
            values = {metric.label: 1.5 for metric in company.metrics}
            path, _ = write_company_workbook(company, values, output_dir=self.out)
            cells = verify_workbook(company, path)
            self.assertEqual(len(cells), 3, msg=company.slug)
            self.assertTrue(all(cell.value == 1.5 for cell in cells), msg=company.slug)

    def test_missing_file_is_reported(self):
        company = find_company("ADI")
        with self.assertRaises(WorkbookError):
            verify_workbook(company, self.out / "nope.xlsx")

    def test_text_in_a_forecast_cell_is_caught(self):
        company = find_company("DE")
        values = {metric.label: 10.0 for metric in company.metrics}
        path, _ = write_company_workbook(company, values, output_dir=self.out)

        workbook = load_workbook(path)
        workbook["Summary"]["C7"] = "12,345"
        workbook.save(path)
        workbook.close()

        with self.assertRaises(WorkbookError) as caught:
            verify_workbook(company, path)
        self.assertIn("must hold a number", str(caught.exception))

    def test_renamed_metric_label_is_caught(self):
        company = find_company("HAS")
        values = {metric.label: 10.0 for metric in company.metrics}
        path, _ = write_company_workbook(company, values, output_dir=self.out)

        workbook = load_workbook(path)
        workbook["Summary"]["A7"] = "Net fee income"
        workbook.save(path)
        workbook.close()

        with self.assertRaises(WorkbookError):
            verify_workbook(company, path)


class HeaderScanTests(unittest.TestCase):
    def test_header_row_is_found_in_every_supplied_template(self):
        for company in load_companies():
            workbook = load_workbook(PATHS.templates / company.output_file)
            try:
                self.assertEqual(
                    find_header_row(workbook["Summary"], company.period), 6, msg=company.slug
                )
            finally:
                workbook.close()

    def test_wrong_period_is_not_matched(self):
        company = find_company("HD")
        workbook = load_workbook(PATHS.templates / company.output_file)
        try:
            with self.assertRaises(WorkbookError):
                find_header_row(workbook["Summary"], "FY2027Q1")
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()
