"""Template-safe workbook writing.

The organisers' validator (``scripts/check-forecasts.mjs``) is unforgiving and
silent about intent: it scans the ``Summary`` sheet for a ``Metric | Units |
<period>`` header row, then requires the next three rows to carry the exact
metric labels and units, with a genuine number in the period column.

So this module never builds a workbook. It copies the supplied template, locates
the header row by scanning rather than assuming, re-checks that the labels and
units still match ``challenge/companies.json``, and writes only the three
forecast cells. Anything unexpected raises rather than producing a file that
looks fine and fails at upload.
"""

from __future__ import annotations

import math
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook

from .config import PATHS, Company, Metric
from .errors import ForecastSystemError

SUMMARY_SHEET = "Summary"
HEADER_SCAN_ROWS = 30
METRIC_COLUMN = 1
UNITS_COLUMN = 2
FORECAST_COLUMN = 3


class WorkbookError(ForecastSystemError):
    """Raised when a workbook cannot be written exactly as the validator expects."""


@dataclass(frozen=True)
class WrittenCell:
    """One forecast as it was actually written to disk."""

    metric: str
    units: str
    cell: str
    value: float

    def as_dict(self) -> dict[str, Any]:
        return {"metric": self.metric, "units": self.units, "cell": self.cell, "value": self.value}


def _cell_text(sheet: Any, row: int, column: int) -> str:
    value = sheet.cell(row=row, column=column).value
    return "" if value is None else str(value).strip()


def find_header_row(sheet: Any, period: str) -> int:
    """Locate the ``Metric | Units | <period>`` header row.

    Scanned rather than hardcoded: the templates put it on row 6 today, but a
    hardcoded index would fail silently if the organisers reissued a template.
    """
    for row in range(1, HEADER_SCAN_ROWS + 1):
        if (
            _cell_text(sheet, row, METRIC_COLUMN) == "Metric"
            and _cell_text(sheet, row, UNITS_COLUMN) == "Units"
            and _cell_text(sheet, row, FORECAST_COLUMN) == period
        ):
            return row
    raise WorkbookError(
        f"No 'Metric | Units | {period}' header row found in the first "
        f"{HEADER_SCAN_ROWS} rows of the {SUMMARY_SHEET} sheet"
    )


def _coerce_forecast(metric: Metric, value: Any) -> float:
    """Convert a forecast to the native float the validator requires."""
    if isinstance(value, bool) or value is None:
        raise WorkbookError(f"{metric.label}: forecast must be a number, got {value!r}")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise WorkbookError(f"{metric.label}: forecast must be a number, got {value!r}") from exc
    if not math.isfinite(number):
        raise WorkbookError(f"{metric.label}: forecast must be finite, got {number!r}")
    return number


def write_company_workbook(
    company: Company,
    forecasts: Mapping[str, Any],
    *,
    output_dir: Path | None = None,
    template_dir: Path | None = None,
) -> tuple[Path, list[WrittenCell]]:
    """Write one completed workbook from the supplied template.

    ``forecasts`` maps metric label to value. Every metric must be present: a
    missing figure scores the maximum penalty, so an incomplete mapping is an
    error here rather than a blank cell discovered at upload time.
    """
    templates = template_dir or PATHS.templates
    destination_dir = output_dir or PATHS.submission
    template_path = templates / company.output_file
    if not template_path.exists():
        raise WorkbookError(f"Template is missing: {template_path}")

    missing = [metric.label for metric in company.metrics if metric.label not in forecasts]
    if missing:
        raise WorkbookError(f"{company.slug}: no forecast supplied for {missing}")

    destination_dir.mkdir(parents=True, exist_ok=True)
    destination = destination_dir / company.output_file
    shutil.copyfile(template_path, destination)

    workbook = load_workbook(destination)
    if SUMMARY_SHEET not in workbook.sheetnames:
        raise WorkbookError(f"{company.output_file}: {SUMMARY_SHEET} sheet is missing")
    sheet = workbook[SUMMARY_SHEET]

    header_row = find_header_row(sheet, company.period)
    written: list[WrittenCell] = []

    for offset, metric in enumerate(company.metrics, start=1):
        row = header_row + offset
        actual_label = _cell_text(sheet, row, METRIC_COLUMN)
        actual_units = _cell_text(sheet, row, UNITS_COLUMN)
        if actual_label != metric.label:
            raise WorkbookError(
                f"{company.output_file} row {row}: expected metric "
                f"{metric.label!r}, template has {actual_label!r}"
            )
        if actual_units != metric.units:
            raise WorkbookError(
                f"{company.output_file} row {row}: expected units "
                f"{metric.units!r}, template has {actual_units!r}"
            )

        value = _coerce_forecast(metric, forecasts[metric.label])
        cell = sheet.cell(row=row, column=FORECAST_COLUMN)
        cell.value = value
        written.append(
            WrittenCell(
                metric=metric.label,
                units=metric.units,
                cell=f"{cell.column_letter}{row}",
                value=value,
            )
        )

    workbook.save(destination)
    workbook.close()
    return destination, written


def verify_workbook(company: Company, path: Path | None = None) -> list[WrittenCell]:
    """Re-read a written workbook and apply the validator's own rules.

    Run after writing so a malformed file is caught by us, in the run log, rather
    than by the organisers' script minutes before the deadline.
    """
    target = path or (PATHS.submission / company.output_file)
    if not target.exists():
        raise WorkbookError(f"{company.output_file}: file is missing")

    workbook = load_workbook(target, data_only=True)
    try:
        if SUMMARY_SHEET not in workbook.sheetnames:
            raise WorkbookError(f"{company.output_file}: {SUMMARY_SHEET} sheet is missing")
        sheet = workbook[SUMMARY_SHEET]
        header_row = find_header_row(sheet, company.period)

        cells: list[WrittenCell] = []
        for offset, metric in enumerate(company.metrics, start=1):
            row = header_row + offset
            if _cell_text(sheet, row, METRIC_COLUMN) != metric.label:
                raise WorkbookError(f"{company.output_file} row {row}: metric label changed")
            if _cell_text(sheet, row, UNITS_COLUMN) != metric.units:
                raise WorkbookError(f"{company.output_file} row {row}: units changed")

            cell = sheet.cell(row=row, column=FORECAST_COLUMN)
            value = cell.value
            if isinstance(value, bool) or not isinstance(value, (int, float)):
                raise WorkbookError(
                    f"{company.output_file}: {metric.label} must hold a number, "
                    f"found {value!r} ({type(value).__name__})"
                )
            if not math.isfinite(float(value)):
                raise WorkbookError(f"{company.output_file}: {metric.label} is not finite")
            cells.append(
                WrittenCell(
                    metric=metric.label,
                    units=metric.units,
                    cell=f"{cell.column_letter}{row}",
                    value=float(value),
                )
            )
        return cells
    finally:
        workbook.close()
